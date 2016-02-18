#!/usr/bin/env python
#
# (C) 2016 Clemson University
#
# See LICENSE in top-level directory.
#
# File: analytics/scripts/orangefs-purge-logs2df.py
# Author: Jeff Denton
#
from __future__ import print_function
import os
import sys
import pandas as pd
from openpyxl.styles import Font
from openpyxl.cell import get_column_letter

# Note: This script has two dependencies that can be installed via pip:
#    pandas
#    openpyxl

# Returns a list of strings representing the log file paths
def get_log_files_list(log_dir):
    log_files = []
    for f in os.listdir(log_dir):
        if f.endswith(".log"):
            log_files.append(log_dir + os.sep + f)
    return log_files

# Returns a list of dicts representing each parsed log file
def parse_log_files(log_files):
    dict_list = []
    for f in log_files:
        with open(f, 'r') as fh:
            lines = []
            for line in fh:
                if line.startswith('K\t') or line.startswith('R\t'):
                    continue
                lines.append(line.strip())
            if len(lines) == 0:
                continue
            d = {}
            for l in lines:
                k, sep, v = l.partition('\t')
                d[k] = v
            dict_list.append(d)
    return dict_list

# Strip the mount prefix from each user directory to create the user field
# Makes the user column the index.
# Sorts by user
def set_index(df, prefix='/mnt/orangefs/'):
    df['user'] = df.apply(lambda x: x['directory'][len(prefix):], axis=1)
    df.set_index('user', inplace=True)
    df.sort_index(inplace=True)

def convert_columns(df):
    uint64_columns = [
            'current_time',
            'removal_basis_time',
            'finish_time',
            'duration_seconds',
            'removed_bytes',
            'removed_files',
            'failed_removed_bytes',
            'failed_removed_files',
            'kept_bytes',
            'kept_files',
            'directories',
            'symlinks',
            'unknown']

    float_columns = [
            'percent_bytes_removed',
            'percent_files_removed',
            'pre_purge_avg_file_size',
            'post_purge_avg_file_size',
            'purged_avg_file_size']

    bool_columns = ['dry_run', 'purge_success']

    dtype_dict_list = [
            {'dtype': 'uint64', 'column_list': uint64_columns},
            {'dtype': 'float', 'column_list': float_columns},
            {'dtype': 'bool', 'column_list': bool_columns}]

    for ct in dtype_dict_list:
        df[ct['column_list']] = df[ct['column_list']].astype(ct['dtype'])

    # Convert integer timestamps to datetime
    df['current_time'] = pd.to_datetime(df['current_time'],unit='s')
    df['removal_basis_time'] = pd.to_datetime(df['removal_basis_time'],unit='s')
    df['finish_time'] = pd.to_datetime(df['finish_time'],unit='s')

def format_xlsx_sheet(df, sheet):

    column_widths = []
    # Determine the widest cell of the index column (including the index name)
    column_widths.append(max([len(df.index.name),
                             max(map(lambda s: len(str(s)), df.index.values))]))

    # For every DataFrame column calculate the widest cell (including the column header)
    for column in df.columns:
        column_widths.append(max([len(column),
                                  max(map(lambda s: len(str(s)), df[column]))]))

    print('    Calculated+Set column_widths = ', column_widths)
    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    font = Font(name='monospace',
                size = 10)

    ss_num_columns = 1 + len(df.columns) # +1 for the index
    ss_num_rows = 1 + len(df) # +1 for the headers
    selection = 'A1:' + get_column_letter(ss_num_columns) + str(ss_num_rows)
    print('    Cell Selection = ', selection)

    print('    Changing font of every cell in the worksheet to name=\'monospace\', size=10')
    for column in sheet.columns:
        for cell in column:
            cell.font = font

def error(*objs):
    print("ERROR: ", *objs, file=sys.stderr)

if __name__ == '__main__':

    if(len(sys.argv) != 4):
        error("Usage: ", sys.argv[0], '<log_dir> <out_dir> <file_name>\n',
              '\n\twhere:\n',
              '\t\t<log_dir> is the directory containing your log results from orangefs-purge:\n',
              '\t\t<out_dir> is the directory where your generated files will be written\n',
              '\t\t<file_name> is the file name prefix given to the generated .pkl and .xlsx' +
              ' files\n')
        exit(1)

    log_dir = sys.argv[1]
    if not os.access(log_dir, os.R_OK | os.X_OK):
        error("log_dir argument must have both read and execute permissions set!")
        exit(1)

    out_dir = sys.argv[2]
    if not os.access(out_dir, os.R_OK | os.W_OK | os.X_OK):
        error("out_dir argument must have: read, write, and execute permissions set!")
        exit(1)

    pkl_out = out_dir + os.sep + sys.argv[3] + '.pkl'
    xlsx_out = out_dir + os.sep + sys.argv[3] + '.xlsx'

    if os.access(pkl_out, os.F_OK) or os.access(xlsx_out, os.F_OK):
        error('One or more of the files to be generated already exists!\n'
              '\tExiting early to avoid overwriting existing file(s)!\n')
        exit(1)

    log_files = get_log_files_list(log_dir)

    if(len(log_files) == 0):
        error("No log files found! Nothing to do!")
        exit(1)

    dict_list = parse_log_files(log_files)

    df = pd.DataFrame()
    df = df.append(dict_list, ignore_index=True)
    print('\nData frame as parsed from the individual log files:\n')
    df.info()
    print('\n\n')

    print('Dropping redundant columns since calculating this information is possible using the\n',
          '    integer type:\n')
    df.drop('current_time_str', axis=1, inplace=True)
    df.drop('removal_basis_time_str', axis=1, inplace=True)
    df.drop('finish_time_str', axis=1, inplace=True)
    df.info()
    print('\n\n')

    print('Computing the user field from the directory field and setting the user field as the\n',
          '    index of the DataFrame:\n')
    set_index(df)
    df.info()
    print('\n\n')

    print('Dropping the directory since the user field is the index now:\n')
    df.drop('directory', axis=1, inplace=True)
    df.info()
    print('\n\n')

    print('Converting all column data types to their proper data type:\n')
    convert_columns(df)
    df.info()
    print('\n\n')

    print('Calculating total_files and total_bytes:\n')
    df['total_files'] = df['removed_files'] + df['failed_removed_files'] + df['kept_files']
    df['total_bytes'] = df['removed_bytes'] + df['failed_removed_bytes'] + df['kept_bytes']
    df.info()
    print('\n\n')

    print('Ordering the columns (this is the final DataFrame state):\n')
    df = df[[
        'duration_seconds',
        'removed_files',
        'removed_bytes',
        'kept_files',
        'kept_bytes',
        'failed_removed_files',
        'failed_removed_bytes',
        'total_files',
        'total_bytes',
        'directories',
        'symlinks',
        'unknown',
        'percent_files_removed',
        'percent_bytes_removed',
        'pre_purge_avg_file_size',
        'purged_avg_file_size',
        'post_purge_avg_file_size',
        'removal_basis_time',
        'current_time',
        'finish_time',
        'purge_success',
        'dry_run'
        ]]
    df.info()
    print('\n\n')

    print('Writing DataFrame to Pickle file:\n')
    # Write df out in both pickle and xlsx format
    df.to_pickle(pkl_out)
    print('    DONE!\n\n')

    print('Preparing .xlsx document:\n')
    writer = pd.ExcelWriter(xlsx_out)
    df.to_excel(writer, 'results')
    # Auto-configure the column width and set monospace font
    format_xlsx_sheet(df, writer.sheets['results'])
    print('\n    DONE!\n\n')

    print('Writing DataFrame to .xlsx format:\n')
    writer.save()
    print('    DONE!\n\n')
